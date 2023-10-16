'''
@Author 浊玉
@Description
@Date create in 2023/6/29 22:25
'''


liucheng = """执行流程：
1.先将房屋信息表放到xlsx第一个。然后才能运行程序！
输入拆分并去重的文件名，如：21号履祥楼
2.等待程序运行。
3.程序运行后，生成newfile.xlsx
#文件说明：updated_file.xlsx  :拆分后的文件，但是未去重
          newfile.xlsx :去重后文件，完成全部操作。
          
          """
print(liucheng)
import openpyxl
import pandas as pd
from tqdm import tqdm

filename = input("请输入文件名，不需要xlsx后缀：")
filename = './'+filename + '.xlsx'
print("被执行文件为：{}".format(filename))
# 打开Excel文件
workbook = openpyxl.load_workbook(filename)

# 选择要操作的工作表
worksheet = workbook.active

# 获取最大行数和最大列数
max_row = worksheet.max_row
max_column = worksheet.max_column

# 新建一个列表来存储更新后的行数据
updated_rows = []

# 遍历每一行，使用tqdm创建进度条
for row_index in tqdm(range(1, 1000 + 1), desc="进度", unit="行"):
    # 获取当前行的数据
    row_data = []
    for column_index in range(1, max_column + 1):
        cell_value = worksheet.cell(row=row_index, column=column_index).value
        row_data.append(cell_value)

    # 如果当前行为空行，停止程序
    if not any(row_data):
        break

    # 获取第9列的值（客户姓名）
    customer_names = row_data[8]  # 列索引从0开始，所以第9列的索引为8

    # 如果客户姓名中包含空格，则分割成多个姓名
    if customer_names and ' ' in customer_names:
        names = customer_names.split(' ')
        # 更新当前行的客户姓名列为第一个姓名
        row_data[8] = names[0].strip()
        updated_rows.append(row_data)

        # 创建新行，并将其余信息保持不变
        for name in names[1:]:
            new_row_data = row_data.copy()
            new_row_data[8] = name.strip()
            updated_rows.append(new_row_data)
    else:
        updated_rows.append(row_data)

# 清空工作表的数据
worksheet.delete_rows(1, max_row)


# 将更新后的行数据写入工作表
for row_data in updated_rows:
    worksheet.append(row_data)

# 保存修改后的Excel文件
workbook.save('updated_file.xlsx')
print("拆分姓名成功！，原始未拆分行数为进度条1000前的行数")

# 读取Excel文件
df = pd.read_excel('updated_file.xlsx')

# 获取第一个工作表
worksheet = df.iloc[:, :].values.tolist()

# 获取总行数
total_rows = len(worksheet)


# 新建一个列表来存储更新后的行数据
updated_rows = []

# 用于跟踪已经出现的用户姓名和第6列数据的组合
seen_combinations = set()

# 遍历每一行，使用tqdm创建进度条
for row_index, row in tqdm(enumerate(worksheet), desc="进度", total=total_rows):
    # 获取第9列的用户姓名和第6列的数据
    name = row[8]  # 列索引从0开始，所以第9列的索引为8
    column6 = row[5]  # 列索引从0开始，所以第6列的索引为5

    # 判断是否同时不同于之前的组合
    if (name, column6) not in seen_combinations:
        updated_rows.append(row)
        seen_combinations.add((name, column6))

# 将更新后的行数据转换为DataFrame
updated_df = pd.DataFrame(updated_rows, columns=df.columns)

# 保存修改后的Excel文件
updated_df.to_excel('newfile.xlsx', index=False)
print("去重成功！拆分后行数为：进度条行数，去重后行数见newfile.xlsx")



